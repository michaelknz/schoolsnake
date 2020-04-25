from PIL import Image, ImageTk
import tkinter
from openpyxl import load_workbook

class Leader_Table:
    def __init__(self,screen):
        self.scrn=screen
        self.Enter_Name()
        self.wb=load_workbook('./leader_table.xlsx')
        self.table=self.wb['1']
        self.led_col=5

    def Enter_Name(self):
        self.scrn.Destroy_Canvas()
        self.scrn.Create_Canvas()
        self.Is_Over=False
        self.pilImage = Image.open("snake.jpg", "r")
        self.img = ImageTk.PhotoImage(self.pilImage)
        self.background_label = tkinter.Label(self.scrn.root, image=self.img)
        self.background_label.place(x=0, y=0, relwidth=1, relheight=1)
        self.name=tkinter.StringVar()
        self.name_enter=tkinter.Entry(textvariable=self.name)
        self.name_enter.place(relx=.5, rely=.4, anchor="c")
        self.button=tkinter.Button(text="Start", background="brown", foreground="white", height=1, width=50, command=self.Button_Event)
        self.button.pack()
        self.scrn.canvas.create_window((self.scrn.width // 2 - 170, self.scrn.height // 2 - 20), anchor="nw", window=self.button)
        self.Loop()
        self.name_enter.destroy()
        self.scrn.Destroy_Canvas()

    def Button_Event(self):
        self.Is_Over=True

    def Loop(self):
        while(True):
            if self.Is_Over:
                break
            self.scrn.root.update()

    def Check_Results(self,mas,score):
        b=True
        for i in range(len(mas)):
            if mas[i][0]==self.name.get():
                b = False
                if score>mas[i][1]:
                    mas[i][1]=score
                    mas.sort(key=lambda x: x[1])
                    mas.reverse()
                    return
        if b:
            mas.append([self.name.get(),score])
        mas.sort(key=lambda x: x[1])
        mas.reverse()
        mas=mas[:self.led_col:]


    def Update_T(self,score):
        mas=[]
        for i in range(1,self.led_col+1):
            if self.table['A'+str(i)].value==None:
                break
            mas.append([self.table['A'+str(i)].value, self.table['B'+str(i)].value])
        self.Check_Results(mas,score)
        for i in range(1,len(mas)+1):
            self.table.cell(row=i, column=1, value=mas[i-1][0])
            self.table.cell(row=i, column=2, value=mas[i-1][1])
        self.wb.save('./leader_table.xlsx')